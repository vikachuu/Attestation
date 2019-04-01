import io
from flask import send_file
from web import db
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side


class FiveYearsPlanUtils:
    @staticmethod
    def get_five_years_plan_document():
        sql = """
        SELECT surname AS "прізвище", name AS "ім'я", middle_name AS "по-батькові", qualification_category AS "категорія", 
        rank AS "звання", previous_attestation_date AS "дата попередньої атсетації", 
        next_attestation_date AS "дата наступної атестації"
        FROM teacher;
        """
        res = db.engine.execute(sql)
        result = res.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Перспективний план"

        # check for current educational year
        start_year = datetime.now().year if datetime.now().month > 5 else datetime.now().year - 1

        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 50

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 17
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 27
        ws.column_dimensions['E'].width = 10

        # make first header
        ws.merge_cells('A1:K1')
        header1_cell = ws['A1']
        header1_cell.value = "«ЗАТВЕРДЖЕНО»\nДиректор спеціалізованої школи №173 Мазаєва Л.А.\n"
        header1_cell.alignment = Alignment(horizontal="right", vertical="center", wrapText=True)

        # make second header
        ws.merge_cells('A2:K2')
        header2_cell = ws['A2']
        header2_cell.value = "Перспективний план курсової перепідготовки та атестації педагогічних працівників " \
                             "спеціалізованої школи №173 на {}-{} н.р.\n".format(start_year, start_year + 5)
        header2_cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        # number cell
        ws.merge_cells('A3:A4')
        number_cell = ws['A3']
        number_cell.value = '№'

        # pib cell
        ws.merge_cells('B3:B4')
        pib_cell = ws['B3']
        pib_cell.value = 'Прізвище\nім\'я\nпо-батькові'

        # attestation cells
        ws.merge_cells('C3:D3')
        prev_cell = ws['C3']
        prev_cell.value = 'Попередня\nатестація'

        next_cell = ws['E3']
        next_cell.value = 'Наступна\nатестація'

        prev_date_cell = ws['C4']
        prev_date_cell.value = 'Рік'

        res_cell = ws['D4']
        res_cell.value = 'Результат'

        next_date_cell = ws['E4']
        next_date_cell.value = 'Рік'

        # perspective plan
        ws.merge_cells('F3:K3')
        plan_cell = ws['F3']
        plan_cell.value = 'Перспективний план атестації'

        year1_cell = ws['F4']
        year1_cell.value = '{}\n—\n{}'.format(start_year, start_year + 1)

        year2_cell = ws['G4']
        year2_cell.value = '{}\n—\n{}'.format(start_year + 1, start_year + 2)

        year3_cell = ws['H4']
        year3_cell.value = '{}\n—\n{}'.format(start_year + 2, start_year + 3)

        year4_cell = ws['I4']
        year4_cell.value = '{}\n—\n{}'.format(start_year + 3, start_year + 4)

        year5_cell = ws['J4']
        year5_cell.value = '{}\n—\n{}'.format(start_year + 4, start_year + 5)

        sign_cell = ws['K4']
        sign_cell.value = 'Підпис'

        # one teacher row
        j = 5
        for i, row in enumerate(result, 1):
            ws['A{}'.format(j)].value = i
            ws['B{}'.format(j)].value = "\n".join(row[:3])
            ws['C{}'.format(j)].value = row[5]
            ws['D{}'.format(j)].value = "категорія \"{}\"\nзвання \"{}\"".format(row[3], row[4]) if row[4] \
                else "категорія \"{}\"".format(row[3])
            ws['E{}'.format(j)].value = row[6]

            # perspective plan
            next_att = row[6]
            a, c = '', ''
            if next_att == datetime.now().year:
                a, c = 'F{}'.format(j), 'J{}'.format(j)  # check if next education year
            elif next_att == datetime.now().year + 1:
                a, c = ('F{}'.format(j), 'J{}'.format(j)) if datetime.now().month > 5 \
                    else ('G{}'.format(j), 'F{}'.format(j))
            elif next_att == datetime.now().year + 2:
                a, c = ('G{}'.format(j), 'F{}'.format(j)) if datetime.now().month > 5 \
                    else ('H{}'.format(j), 'G{}'.format(j))
            elif next_att == datetime.now().year + 3:
                a, c = ('H{}'.format(j), 'G{}'.format(j)) if datetime.now().month > 5 \
                    else ('I{}'.format(j), 'H{}'.format(j))
            elif next_att == datetime.now().year + 4:
                a, c = ('I{}'.format(j), 'H{}'.format(j)) if datetime.now().month > 5 \
                    else ('J{}'.format(j), 'I{}'.format(j))
            elif next_att == datetime.now().year + 5 and datetime.now().month > 5:
                a, c = ('J{}'.format(j), 'I{}'.format(j))

            if a and c:  # in 5 years
                ws[a].value = 'A'  # attestation
                ws[c].value = 'K'  # courses

            j += 1

        # set borders for all cells
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows('A3:K{}'.format(len(result) + 4)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         attachment_filename='attestation_plan.xlsx', as_attachment=True)
